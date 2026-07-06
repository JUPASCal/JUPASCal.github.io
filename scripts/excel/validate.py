#!/usr/bin/env python
"""
Validation harness: recalc the generated Excel for each test student and diff
計分版!J against the TS oracle (build/oracle_scores.json from oracle.mjs).

Uses LibreOffice `--convert-to` with a force-recalc-on-load profile (clean, no
lingering UNO server). Grades 1–5 are written as NUMBERS (5*/5**/U as text) to
match the workbook's grade scale. See BUILD_PLAN.md "Validation".

  node scripts/excel/oracle.mjs
  ~/miniconda3/envs/jupascal/bin/python scripts/excel/build_2026_excel.py
  ~/miniconda3/envs/jupascal/bin/python scripts/excel/validate.py   # needs soffice
"""
import os, json, subprocess, shutil, collections, warnings
import openpyxl

warnings.simplefilter("ignore")
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
GEN = os.path.join(ROOT, "build", "2026 JUPAS Cal (generated).xlsx")
RECALC = os.path.join(ROOT, "build", "recalc")
PROFILE = "/tmp/lo_prof"
TOL = 0.1
ELECT_PH = ["請選擇第一選修科", "請選擇第二選修科", "請選擇第三選修科", "請選擇第四選修科"]


def grade_cell(g):
    """1–5 are numbers in the workbook's scale; 5*/5**/U stay text."""
    return int(g) if g in ("1", "2", "3", "4", "5") else g


def seed_profile():
    d = os.path.join(PROFILE, "user")
    os.makedirs(d, exist_ok=True)
    with open(os.path.join(d, "registrymodifications.xcu"), "w") as f:
        f.write(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
            'xmlns:xs="http://www.w3.org/2001/XMLSchema" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
            ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
            '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
            ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
            '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
            '</oor:items>\n')


def write_input(student, path):
    wb = openpyxl.load_workbook(GEN)
    h = wb["主頁"]
    h["D5"] = "否"
    c = student["cores"]
    h["C8"], h["C9"], h["C10"] = grade_cell(c["chi"]), grade_cell(c["eng"]), grade_cell(c["math"])
    h["C11"] = "達標"
    h["C12"] = grade_cell(student["m12"]["grade"]) if student.get("m12") else "請選擇等級"
    for i in range(4):
        h[f"B{15 + i}"], h[f"C{15 + i}"] = ELECT_PH[i], "請選擇等級"
    for i, e in enumerate(student["electives"]):
        h[f"B{15 + i}"], h[f"C{15 + i}"] = e["zh"], grade_cell(e["grade"])
    wb.save(path)


def recalc(inputs):
    env = dict(os.environ)
    env["HOME"] = "/tmp/lohome"          # soffice needs a writable HOME
    env["SAL_USE_VCLPLUGIN"] = "svp"     # headless VCL
    os.makedirs("/tmp/lohome", exist_ok=True)
    out = os.path.join(RECALC, "out")
    os.makedirs(out, exist_ok=True)
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", out, *inputs,
         f"-env:UserInstallation=file://{PROFILE}"],
        env=env, stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL, timeout=300)
    subprocess.run(["pkill", "-9", "-f", "soffice.bin"], stderr=subprocess.DEVNULL)
    return out


def main():
    students = json.load(open(os.path.join(ROOT, "scripts", "excel", "test_students.json")))
    oracle = json.load(open(os.path.join(ROOT, "build", "oracle_scores.json")))
    insts = {p["jupas_code"]: p["institution"]
             for p in json.load(open(os.path.join(ROOT, "data/processed/JUPAS_2026_Unified_Data.json")))}
    shutil.rmtree(RECALC, ignore_errors=True)
    os.makedirs(RECALC)
    seed_profile()
    inputs = []
    for s in students:
        p = os.path.join(RECALC, f"in_{s['name']}.xlsx")
        write_input(s, p)
        inputs.append(p)
    outdir = recalc(inputs)

    rep = open(os.path.join(ROOT, "build", "val_report.txt"), "w", buffering=1)
    def pr(msg):
        print(msg); rep.write(msg + "\n"); rep.flush()

    pr(f"=== Excel vs TS oracle (±{TOL}) ===")
    grand = 0
    for s in students:
        ws = openpyxl.load_workbook(os.path.join(outdir, f"in_{s['name']}.xlsx"),
                                    data_only=True)["計分版"]
        codes = {ws[f"A{r}"].value: ws[f"J{r}"].value
                 for r in range(36, 458) if ws[f"A{r}"].value is not None}
        orc = oracle[s["name"]]
        mism = []
        for k, v in orc.items():
            ex = codes.get(k)
            ex = ex if isinstance(ex, (int, float)) else 0.0
            if abs(ex - v) > TOL:
                mism.append((abs(ex - v), k, round(ex, 2), round(v, 2)))
        mism.sort(reverse=True)
        grand += len(mism)
        by_inst = dict(collections.Counter(insts.get(k) for _, k, _, _ in mism))
        pr(f"[{s['name']:12}] {len(orc) - len(mism)}/{len(orc)} match; "
           f"{len(mism)} mismatch  by-inst={by_inst}")
        for d, k, ex, ts in mism[:6]:
            pr(f"      {k} ({insts.get(k)}): excel={ex} ts={ts} Δ={round(d, 2)}")
    pr(f"total mismatches across {len(students)} students: {grand}")


if __name__ == "__main__":
    main()
