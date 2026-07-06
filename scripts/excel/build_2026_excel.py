#!/usr/bin/env python
"""
2026 JUPAS Cal Excel generator.

Strategy (per the user's decision): **clone the shell, regenerate the guts.**
Load the proven 2025 workbook as a styled template — keeping its appearance,
static sheets, prose, formatting, named ranges and dropdowns — then rebuild only
the per-programme DATA + formula-pattern regions from the unified 2026 dataset.

LOGIC lives in docs/manuals/EXCEL_LOGIC.md (§ references below). DATA comes from
JUPAS_2026_Unified_Data.json via unified_model.py. Re-runnable each cycle.

Run:  ~/miniconda3/envs/jupascal/bin/python scripts/excel/build_2026_excel.py
Output: build/2026 JUPAS Cal (generated).xlsx

STATUS: scaffold. Shell-clone + year-stamp work end-to-end today; each
build_<sheet>() is stubbed and implemented incrementally (see BUILD_PLAN.md).
"""
from __future__ import annotations
import warnings
import re
import openpyxl

from layout import (
    TEMPLATE_XLSX, OUTPUT_XLSX, ENGINE_DATA_START_ROW, INST_ORDER,
    RSC_DATA_START_ROW, ELIG_DATA_START_ROW,
)
from unified_model import load_programmes, group_by_institution, ordered_programmes
from steps import (
    emit_weights, emit_weighted, emit_required, emit_boost, emit_mask,
    emit_aggregate, emit_best_of, emit_max_weighted, scale_row, recipe_exceptions,
    get_target_count, get_bonus, EN_TO_ZH,
)

# Exception features with a real step override implemented (vs default fallback).
HANDLED_EXCEPTIONS = {
    "best_of_weights", "hkust_weighted_best", "bonus_6th", "bonus_7th",
    "additional_bonus_6th", "compulsory_subject_pool", "maths_m1m2_as_one",
    "m1m2_half_replacement", "max_weighted_subjects",
}
import collections

warnings.simplefilter("ignore")  # x14 extension CF/DV warnings — see BUILD_PLAN risks

CYCLE = 2026

# formula_2025_id -> the method label the engine mask matches on (EXCEL_LOGIC §4.4).
# NB: 2026 data uses Best-N only (no 3C2X); 'best6' is new vs the 2025 engine,
# which needs a Best-6 mask branch added in stage 3.
METHOD_LABEL = {"best5": "Best 5", "best4": "Best 4", "best6": "Best 6"}


def _cell(v):
    """Match the workbook convention: '/' for a missing value."""
    return v if v not in (None, "") else "/"


def _req_level(v, default=1):
    """Parse a requirement level ('3', '4', '5^' ...) to int; blank -> default."""
    if v in (None, ""):
        return default
    digits = "".join(ch for ch in str(v) if ch.isdigit())
    return int(digits) if digits else default


def _apl_accepts(policy):
    """True if the programme accepts ApL (apl_policy is anything but 'none')."""
    return policy not in (None, "none", "", [])


def assign_engine_rows(groups):
    """Compute the 計分版 start row of each institution block (contiguous, +2
    spacer rows for the per-institution header label, matching the 2025 layout)."""
    rows, r = {}, ENGINE_DATA_START_ROW
    for inst in INST_ORDER:
        rows[inst] = r
        r += len(groups[inst]) + 2
    return rows


# --- per-sheet rebuilders (stubs; see EXCEL_LOGIC.md § in each docstring) -----

def stamp_metadata(wb, progs):
    """主頁 title/version + a year-label sweep on the visible UI sheets. Scores use
    Y-1 (2025) benchmarks and 2026 requirements, so labels shift 2024→2025 and
    2025→2026 (2025→2026 first to avoid double-shift). Only non-formula string
    labels on 主頁 / All in One are touched — never data sheets (RSC remarks etc.)."""
    wb["主頁"]["B2"] = f"{CYCLE} JUPAS 計分器 (generated)"
    subs = [("2025", "2026"), ("2024", "2025"), ("25年", "26年"), ("24年", "25年")]
    n = 0
    for name in ("主頁", "All in One"):
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and not v.startswith("=") and type(cell).__name__ != "MergedCell":
                    nv = v
                    for a, b in subs:
                        nv = nv.replace(a, b)
                    if nv != v:
                        cell.value = nv
                        n += 1
    print(f"  [stamp] 主頁!B2 + {n} year labels swept on 主頁/All in One")


def patch_shell(wb):
    """Fix defects in the template's kept header/conversion zone (rows 1-35).
    - HKMU scale row 11 is missing its M1/2 conversion (O11 empty) → M1/2 never
      scored for HKMU. HKMU shares row 13's linear scale, so copy O13's formula.
    """
    eng = wb["計分版"]
    if eng["O11"].value is None:
        eng["O11"] = eng["O13"].value
        print("  [patch] 計分版!O11 (HKMU M1/2 conversion) filled from O13")
    # M1/M2 module selector (D12) — needed for programmes weighting M1 ≠ M2.
    from openpyxl.worksheet.datavalidation import DataValidation
    home = wb["主頁"]
    home["D12"] = "M1"
    home["D11"] = "M1定M2?"
    dv = DataValidation(type="list", formula1='"M1,M2"', allow_blank=True)
    dv.add("D12")
    home.add_data_validation(dv)
    print("  [patch] 主頁!D12 M1/M2 module selector added (default M1)")


def patch_elective_dropdowns(wb):
    """Repair + extend the 選單 elective picker (the elective dropdown machine).

    The 主頁 elective slots (B15:B18) each pick from a named range. The list source
    is 選單!B3:B_n (the master subject list); columns C/D/E/F are per-slot de-dup
    MIRRORS — each blanks a subject already chosen in the OTHER three slots (off the
    live picks at 選單!B29:B32 = 主頁!B15:B18), so the same elective can't be picked
    twice. Two defects ship in the legacy shell, plus a data change:
      • 第一選修科 (slot 1) points at #REF!  → no dropdown.
      • 第四選修科 (slot 4) points at F2:F6 — a stray 4-subject hand list.
      • The two Technology and Living STRANDS are distinct scored subjects
        (steps.EN_TO_ZH) but the picker listed a single ambiguous 科技與生活.
    Fix: split T&L into its two strands in the master list, rebuild C/D/E/F fully
    over the master range, repoint all four named ranges, and wire the slot-1/slot-3
    data validations the template dropped."""
    from openpyxl.worksheet.datavalidation import DataValidation
    s = wb["選單"]
    tl_food = EN_TO_ZH["Technology and Living (Food Science and Technology)"]
    tl_fashion = EN_TO_ZH["Technology and Living (Fashion, Clothing and Textiles)"]
    # 1. master list: rename the ambiguous 科技與生活 → Food strand, append Fashion.
    if s["B21"].value == "科技與生活":
        s["B21"] = tl_food
    s["B26"] = tl_fashion
    last = 26                       # master list now spans B3:B26 (24 subjects)
    # 2. per-slot de-dup mirrors. Each slot hides subjects picked in the other three
    #    (picks: B29..B32 = slots 1..4). Row 2 = the "請選擇…" placeholder.
    others = {"C": ("$B$30", "$B$31", "$B$32"),   # slot 1 hides 2,3,4
              "D": ("$B$29", "$B$31", "$B$32"),   # slot 2 hides 1,3,4
              "E": ("$B$29", "$B$30", "$B$32"),   # slot 3 hides 1,2,4
              "F": ("$B$29", "$B$30", "$B$31")}   # slot 4 hides 1,2,3
    placeholder = {"C": "請選擇第一選修科", "D": "請選擇第二選修科",
                   "E": "請選擇第三選修科", "F": "請選擇第四選修科"}
    for col, (a, b, c) in others.items():
        s[f"{col}2"] = placeholder[col]
        for r in range(3, last + 1):
            s[f"{col}{r}"] = f'=IF(OR({a}=B{r},{b}=B{r},{c}=B{r}),"",B{r})'
    # 3. repoint the four elective named ranges to the full, repaired lists.
    for nm, col in (("第一選修科", "C"), ("第二選修科", "D"),
                    ("第三選修科", "E"), ("第四選修科", "F")):
        wb.defined_names[nm].value = f"選單!${col}$2:${col}${last}"
    # 4. wire the slot-1/slot-3 dropdowns the template dropped (slots 2 & 4 kept).
    home = wb["主頁"]
    for cell, nm in (("B15", "第一選修科"), ("B17", "第三選修科")):
        dv = DataValidation(type="list", formula1=nm, allow_blank=True)
        dv.add(cell)
        home.add_data_validation(dv)
    print("  [dropdown] 選單 electives repaired: slot-1/4 lists fixed, "
          "T&L split into 2 strands, all 4 slot dropdowns wired")


def build_reference_score_calculation(wb, ordered):
    """RSC master DB rows (EXCEL_LOGIC §6), dense layout starting row 13.

    Stage 1 writes the literal identity + benchmark + quota columns straight from
    the unified data (Year-Labeling: benchmarks = scores_2025). Deferred to later
    stages: V/W/X/Y offer formulas (→ Offer Statistics, stage 5), P/Q/R + AH/AI/AJ
    weighting tiers (→ engine, stage 3), AA/AB requirement mirrors (→ §7), and the
    K/L/M reverse-engineering machinery (dropped — benchmarks come from the data).
    """
    ws = wb["Reference Score Calculation"]
    # Clean slate: clear the old 2025 data region (rows 13+), preserving the
    # header (row 12) and the All-in-One lookup rows (3-9) above it.
    for r in range(RSC_DATA_START_ROW, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None
    for i, p in enumerate(ordered):
        r = RSC_DATA_START_ROW + i
        s = p.benchmarks
        ws[f"B{r}"] = p.code
        ws[f"C{r}"] = p.inst_xlsx
        ws[f"D{r}"] = _cell(p.raw.get("faculty"))
        ws[f"E{r}"] = p.name_zh or p.name_en
        ws[f"F{r}"] = METHOD_LABEL.get(p.method_id, p.method_id)
        ws[f"G{r}"] = _cell(s.get("mean"))
        ws[f"H{r}"] = _cell(s.get("uq"))
        ws[f"I{r}"] = _cell(s.get("median"))
        ws[f"J{r}"] = _cell(s.get("lq"))
        ws[f"U{r}"] = _cell(p.quota)
        # offer/intake/competition from the latest year's offer_statistics
        offers = p.offer_statistics or []
        apps = {e["Year"]: e for e in offers if e.get("Type") == "Application"}
        offs = {e["Year"]: e for e in offers if e.get("Type") == "Offer"}
        if apps:
            ws[f"W{r}"] = _cell(apps[max(apps)].get("Band A"))         # Band-A applications
        if offs:
            y = max(offs)
            ws[f"X{r}"] = _cell(offs[y].get("Band A"))                 # Band-A offers
            ws[f"V{r}"] = _cell(offs[y].get("Total"))                  # intake (admitted)
        ws[f"Y{r}"] = f'=IFERROR(W{r}/X{r},"新科目")'                    # competition ratio
    print(f"  [RSC] wrote {len(ordered)} rows (13-{12 + len(ordered)}); "
          f"identity/benchmarks/quota + offer V/W/X/Y")


def build_engine(wb, ordered):
    """計分版 per-programme rows (EXCEL_LOGIC §4.4-4.8), dense from row 36, via the
    composable step pipeline (steps.py). Header/conversion zone rows 1-35 (scales,
    Cat-C matrix, All-in-One block) are kept shell. Cross-sheet offsets:
    計分版 row = RSC row + 23 = 入學要求 row + 14.

    STAGE 3 (default pipeline): convert→weight→select→aggregate for the common
    Best-N case. Exception steps (best_of_weights, HKUST, bonuses, half-
    replacement) currently fall back to default and are logged for follow-up.
    """
    ws = wb["計分版"]
    for r in range(ENGINE_DATA_START_ROW, ws.max_row + 1):  # clear programme block
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None
    warns, exceptions = collections.Counter(), collections.Counter()
    def warn(m): warns[m] += 1
    for i, p in enumerate(ordered):
        r = ENGINE_DATA_START_ROW + i
        rsc, elig = r - 23, r - 14
        # identity + cross-sheet links
        ws[f"A{r}"] = p.code
        ws[f"B{r}"] = f"='Reference Score Calculation'!B{rsc}"
        ws[f"C{r}"] = f"='Reference Score Calculation'!E{rsc}"
        ws[f"E{r}"] = f"='Reference Score Calculation'!F{rsc}"
        ws[f"G{r}"] = f"=入學要求!AB{elig}"   # M1/2 gate
        ws[f"H{r}"] = f"=入學要求!AC{elig}"   # Cat-C/language gate
        ws[f"I{r}"] = f"=入學要求!AD{elig}"   # ApL gate (informational)
        # step pipeline (default + implemented overrides)
        sr = scale_row(p)
        emit_weights(ws, r, p, warn)
        if p.raw.get("best_of_weights_2025"):        # weight-step override
            emit_best_of(ws, r, p, sr, warn)
        emit_weighted(ws, r, sr, p)
        mw = next((c for c in (p.raw.get("calculation_constraints") or [])
                   if isinstance(c, dict) and c.get("type") == "max_weighted_subjects"), None)
        if mw:                                        # cap weighted subjects
            emit_max_weighted(ws, r, p, sr, int(mw.get("limit", 3)))
        emit_required(ws, r, p, sr)
        half_replace = any(c.get("type") == "m1m2_half_replacement"
                           for c in (p.raw.get("calculation_constraints") or []) if isinstance(c, dict))
        emit_boost(ws, r, half_replace)
        n = get_target_count(p)
        emit_mask(ws, r, n, half_replace)
        mode, rate = get_bonus(p)
        emit_aggregate(ws, r, n, sr, mode, rate, half_replace)
        for ex in recipe_exceptions(p):
            if ex not in HANDLED_EXCEPTIONS:
                exceptions[ex] += 1
    print(f"  [計分版] wrote {len(ordered)} engine rows (36-{35 + len(ordered)}) "
          f"via default step pipeline")
    if warns:
        print("    ⚠ EN→ZH coverage gaps:", dict(warns))
    if exceptions:
        tot = sum(exceptions.values())
        print(f"    ⚠ {tot} programme-features on default fallback (need exception steps):")
        for k, v in exceptions.most_common():
            print(f"        {v:4}  {k}")


def build_eligibility(wb, ordered):
    """入學要求 rows (EXCEL_LOGIC §7), dense layout starting row 22.

    Requirement levels L:Q from min_requirements_2026 (current cycle), the check
    formulas S:X against the student-grade row 3 (kept shell), verdict Z, and the
    gate flags AB/AC/AD that feed the engine's G/H/I. The header/shell rows 1-21
    (incl. the student-grade row 3) are preserved.

    NOTE (matches EXCEL_LOGIC §7.3): elective checks are LEVEL-ONLY (compare best
    electives vs required level), NOT subject-specific — faithful to the Excel's
    known limitation. AB/AC default to the dominant 2025 value ('v'/'E'); the ~11
    M1/2 exclusions have no clean unified signal yet.
    """
    ws = wb["入學要求"]
    for r in range(ELIG_DATA_START_ROW, ws.max_row + 1):  # clear old data rows
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None
    for i, p in enumerate(ordered):
        r = ELIG_DATA_START_ROW + i
        rsc = r - 9  # RSC row for this programme (入學 22 <-> RSC 13)
        req = p.requirements
        e1 = req.get("elect1") or {}
        e2 = req.get("elect2") or {}
        # identity — mirror RSC so names/quota stay in sync
        ws[f"A{r}"] = f"='Reference Score Calculation'!B{rsc}"
        ws[f"B{r}"] = f"='Reference Score Calculation'!B{rsc}"
        ws[f"C{r}"] = f"='Reference Score Calculation'!E{rsc}"
        ws[f"D{r}"] = f"='Reference Score Calculation'!U{rsc}"
        # requirement levels (L=chi M=eng N=math O=csd P=E1 Q=E2)
        ws[f"L{r}"] = _req_level(req.get("chi"))
        ws[f"M{r}"] = _req_level(req.get("eng"))
        ws[f"N{r}"] = _req_level(req.get("math"))
        ws[f"O{r}"] = 1  # CSD 達標 required (csd='A' for all programmes)
        ws[f"P{r}"] = _req_level(e1.get("grade"))
        ws[f"Q{r}"] = _req_level(e2.get("grade"))
        # pass/fail checks (cores direct; electives vs student's best/2nd-best)
        ws[f"S{r}"] = f"=IF(L{r}>L$3,0,1)"
        ws[f"T{r}"] = f"=IF(M{r}>M$3,0,1)"
        ws[f"U{r}"] = f"=IF(N{r}>N$3,0,1)"
        ws[f"V{r}"] = f"=IF(O{r}>O$3,0,1)"
        ws[f"W{r}"] = f"=IF(P{r}>LARGE($P$3:$U$3,1),0,1)"
        ws[f"X{r}"] = f"=IF(Q{r}>LARGE($P$3:$U$3,2),0,1)"
        # verdict: AND of all checks (1 eligible / 0 not)
        ws[f"Z{r}"] = f"=S{r}*T{r}*U{r}*V{r}*W{r}*X{r}"
        # gate flags consumed by the engine (G/H/I)
        ws[f"AB{r}"] = "v"   # M1/2 counts (default; ~97% in 2025)
        ws[f"AC{r}"] = "E"   # language slot enabled (default)
        ws[f"AD{r}"] = "v" if _apl_accepts(p.apl_policy) else "NO"
    print(f"  [入學要求] wrote {len(ordered)} rows (22-{21 + len(ordered)}); "
          f"reqs L:Q, checks S:X, verdict Z, gates AB/AC/AD")


def build_resolver(wb, groups):
    """A123 resolver (EXCEL_LOGIC §5). No rebuild needed: the kept shell's mirror
    rows (21+i) reference RSC 13+i / 計分版 J 36+i / 入學要求 Z 22+i — exactly the
    canonical dense offsets this generator uses, so it resolves correctly as-is
    (verified: front-page 主頁!M15 matches 計分版!J for the typed code)."""
    print("  [A123] resolver kept as-is — shell offsets match the dense layout")


def build_data_stores(wb, progs):
    """Offer Statistics + Programme List (EXCEL_LOGIC §8.4) — hidden reference
    sheets. No longer consumed: the competition stat is wired directly into RSC
    V/W/X/Y from each programme's offer_statistics (see build_reference_score_
    calculation). Clear their stale 2025 rows so they don't display outdated data;
    the authoritative source is data/processed/JUPAS_2026_Unified_Data.json.
    (Full multi-year re-population + legacy programmes: deferred follow-up.)"""
    for name in ("Offer Statistics", "Programme List (2024 & 2025)"):
        ws = wb[name]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None and type(cell).__name__ != "MergedCell":
                    cell.value = None
    print("  [data] Offer Statistics / Programme List — stale rows cleared "
          "(reference-only; source = unified JSON)")


PER_SCHOOL_SHEETS = {
    "CityU": "CityUHK", "HKBU": "HKBU", "PolyU": "PolyU", "CUHK": "CUHK",
    "HKUST": "HKUST", "HKU": "HKU", "LingU": "LingnanU", "EdUHK": "EdUHK",
    "HKMU": "HKMU", "SSSDP": "SSSDP",
}


def _remap_row(f, p, r, old_eng, new_eng):
    """Rewrite a per-school row-2 template formula for data row r (position p) of
    the institution's block: cross-sheet row refs shift to the new block, self
    refs (row 2) shift to r. Offsets: RSC = eng-23, 入學要求 = eng-14."""
    if not isinstance(f, str) or not f.startswith("="):
        return f
    old_rsc, old_elig = old_eng - 23, old_eng - 14
    rsc, eng, elig = (new_eng - 23) + p, new_eng + p, (new_eng - 14) + p
    f = f.replace(f"計分版!J{old_eng}", f"計分版!J{eng}")
    f = re.sub(rf"('Reference Score Calculation'!\$?[A-Z]{{1,2}}\$?){old_rsc}(?!\d)",
               lambda m: m.group(1) + str(rsc), f)
    f = re.sub(rf"(入學要求!\$?[A-Z]{{1,2}}\$?){old_elig}(?!\d)",
               lambda m: m.group(1) + str(elig), f)
    # self refs: a bare col+row-2 (not $-absolute, not part of a longer number)
    f = re.sub(r"(?<![\$\d!])([A-Z]{1,2})2(?!\d)", lambda m: m.group(1) + str(r), f)
    return f


def build_per_school_sheets(wb, ordered):
    """Per-institution display sheets (EXCEL_LOGIC §8.5). Each sheet's row-2
    formulas are a template of row-based refs into RSC/計分版/入學要求; regenerate
    the data rows against each institution's NEW (dense) block."""
    starts, counts = {}, {}
    for i, p in enumerate(ordered):
        starts.setdefault(p.inst_xlsx, ENGINE_DATA_START_ROW + i)
        counts[p.inst_xlsx] = counts.get(p.inst_xlsx, 0) + 1
    for sheet in PER_SCHOOL_SHEETS:
        ws = wb[sheet]
        old_eng = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=c).value
            if isinstance(v, str):
                m = re.search(r"計分版!J(\d+)", v)
                if m:
                    old_eng = int(m.group(1)); break
        if old_eng is None:
            continue
        template = {c: ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)}
        new_eng, count = starts[sheet], counts[sheet]

        def put(r, c, val):  # skip merged (non-anchor) cells — their value is read-only
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != "MergedCell":
                cell.value = val

        for r in range(2, ws.max_row + 1):          # clear old data rows (keep header)
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value is not None:
                    put(r, c, None)
        for pos in range(count):
            r = 2 + pos
            for c, f in template.items():
                nf = _remap_row(f, pos, r, old_eng, new_eng)
                if nf is not None:
                    put(r, c, nf)
    print(f"  [per-school] rebuilt {len(PER_SCHOOL_SHEETS)} institution sheets")


# --- orchestrator ------------------------------------------------------------

def main():
    print(f"Loading template: {TEMPLATE_XLSX}")
    wb = openpyxl.load_workbook(TEMPLATE_XLSX, data_only=False)

    progs = load_programmes()
    groups = group_by_institution(progs)
    ordered = ordered_programmes(progs)
    engine_rows = assign_engine_rows(groups)
    print(f"Loaded {len(progs)} programmes across {len(INST_ORDER)} institutions")
    for inst in INST_ORDER:
        print(f"    {inst:8} {len(groups[inst]):>4}  engine block @ row {engine_rows[inst]}")

    print("Rebuilding regions:")
    stamp_metadata(wb, progs)
    patch_shell(wb)
    patch_elective_dropdowns(wb)
    build_reference_score_calculation(wb, ordered)
    build_engine(wb, ordered)
    build_eligibility(wb, ordered)
    build_resolver(wb, groups)
    build_data_stores(wb, progs)
    build_per_school_sheets(wb, ordered)

    wb.save(OUTPUT_XLSX)
    print(f"Saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
